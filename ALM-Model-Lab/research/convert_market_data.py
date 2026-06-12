"""
Converter: SOFR_Market_Data_<YYYYMMDD>.xlsx -> public/market_<YYYY-MM-DD>.json

Reads the 4-sheet BBG market-data workbook (SOFR_OIS_Curve, FHLB_Curve,
Cap_Volatility, ATM_Swaption_Volatility) and emits the lab's market snapshot
JSON. This is the single source of truth for the snapshot contract; the
in-browser xlsx importer (ImportTab) must produce identical JSON.

Conventions enforced here (must match src/math/rates/marketData.ts):
  - rates and vols in DECIMAL per annum, rounded to 1e-7;
  - tenor labels normalized: '1 D' -> '1D', '1 MO' -> '1M', '1 YR' -> '1Y',
    '1Yr' -> '1Y', '1Mo' -> '1M', '18Mo' -> '18M';
  - t_years: 1D = 1/360, k months = k/12 (rounded 1e-6), k years = k;
  - cap strikes are ABSOLUTE RATES (strike_axis = 'absolute_rate'); the
    column labels -2.00%..7.00% are strike levels (negative strikes are a
    negative-rate-era BBG matrix convention), and the explicit 'ATM' column
    is a separate ATM-strike quote, NOT the 0.00% strike column. Verified
    across the 2025-09-30 and 2026-03-31 vintages: the smile trough sits at
    the absolute forward level (~4%) in both, and the ATM quote matches the
    trough column, which rules out a moneyness-offset axis;
  - Term LP = FHLB - SOFR per tenor, 1D pinned to 0 by convention (the bank
    funds overnight at policy rates, not via FHLB advances); the 1D TLP node
    is written at t_years = 0 to anchor the interpolation.

Run:  py research/convert_market_data.py
      py research/convert_market_data.py --src <xlsx> --date 2025-09-30
"""
from __future__ import annotations

import argparse
import json
import os
import re

import openpyxl

HERE = os.path.dirname(__file__)
DEFAULT_SRC = os.path.join(HERE, "data", "SOFR_Market_Data_20260331.xlsx")
DEFAULT_DATE = "2026-03-31"

SHEETS = ["SOFR_OIS_Curve", "FHLB_Curve", "Cap_Volatility", "ATM_Swaption_Volatility"]


def norm_term(label: str) -> str:
    """'1 D' -> '1D', '1 MO' -> '1M', '10 YR' -> '10Y', '1Yr' -> '1Y', '18Mo' -> '18M'."""
    m = re.fullmatch(r"\s*(\d+)\s*(D|MO|YR|Mo|Yr)\s*", str(label))
    if not m:
        raise ValueError(f"unrecognized tenor label: {label!r}")
    unit = {"D": "D", "MO": "M", "Mo": "M", "YR": "Y", "Yr": "Y"}[m.group(2)]
    return f"{m.group(1)}{unit}"


def t_years(term: str) -> float:
    n, unit = int(term[:-1]), term[-1]
    if unit == "D":
        return round(n / 360.0, 6)
    if unit == "M":
        return round(n / 12.0, 6)
    return float(n)


def dec(value: float) -> float:
    """Round to 1e-7 (0.001 bp) to keep the JSON free of xlsx float noise."""
    return round(float(value), 7)


def strike_key(label: str) -> str:
    """'-2.00%' -> '-0.02', '0.25%' -> '0.0025', 'ATM' -> 'ATM'. Absolute strikes."""
    s = str(label).strip()
    if s.upper() == "ATM":
        return "ATM"
    m = re.fullmatch(r"(-?\d+(?:\.\d+)?)%", s)
    if not m:
        raise ValueError(f"unrecognized cap strike label: {label!r}")
    return repr(round(float(m.group(1)) / 100.0, 6))


def read_rows(wb, name: str) -> list[tuple]:
    if name not in wb.sheetnames:
        raise ValueError(f"missing sheet {name!r}; found {wb.sheetnames}")
    return list(wb[name].iter_rows(values_only=True))


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--src", default=DEFAULT_SRC, help="source xlsx workbook")
    ap.add_argument("--date", default=DEFAULT_DATE, help="calibration date YYYY-MM-DD")
    ap.add_argument("--out", default=None, help="output JSON (default public/market_<date>.json)")
    args = ap.parse_args()
    out_path = args.out or os.path.join(HERE, "..", "public", f"market_{args.date}.json")

    wb = openpyxl.load_workbook(args.src, read_only=True, data_only=True)

    # --- SOFR OIS curve: Term | InstType | Mid ---------------------------------
    rows = read_rows(wb, "SOFR_OIS_Curve")
    assert [str(c).strip() for c in rows[0][:3]] == ["Term", "InstType", "Mid"]
    instruments = []
    for term_raw, inst_type, mid in rows[1:]:
        if term_raw is None:
            continue
        term = norm_term(term_raw)
        if inst_type not in ("CASH", "SWAP"):
            raise ValueError(f"unexpected InstType {inst_type!r} at {term}")
        instruments.append(
            {"term": term, "t_years": t_years(term), "type": inst_type, "rate": dec(mid)}
        )

    # --- FHLB curve: Term | Mid ------------------------------------------------
    rows = read_rows(wb, "FHLB_Curve")
    assert [str(c).strip() for c in rows[0][:2]] == ["Term", "Mid"]
    fhlb = []
    for term_raw, mid in rows[1:]:
        if term_raw is None:
            continue
        term = norm_term(term_raw)
        fhlb.append({"term": term, "t_years": t_years(term), "rate": dec(mid)})

    sofr_by_term = {q["term"]: q["rate"] for q in instruments}
    if [q["term"] for q in fhlb] != [q["term"] for q in instruments]:
        raise ValueError("FHLB tenor set does not match SOFR curve tenor set")

    # --- Term LP = FHLB - SOFR, 1D pinned to 0 at t_years = 0 ------------------
    tlp_nodes = []
    for q in fhlb:
        if q["term"] == "1D":
            tlp_nodes.append({"term": "1D", "t_years": 0.0, "spread": 0.0})
            continue
        spread = dec(q["rate"] - sofr_by_term[q["term"]])
        if spread <= 0:
            raise ValueError(f"non-positive Term LP at {q['term']}: {spread}")
        tlp_nodes.append({"term": q["term"], "t_years": q["t_years"], "spread": spread})

    # --- Cap vol surface: banner row, header row, then expiry rows -------------
    rows = read_rows(wb, "Cap_Volatility")
    header = rows[1]
    assert str(header[0]).strip() == "Expiry"
    strike_keys = [strike_key(c) for c in header[1:] if c is not None]
    if "ATM" not in strike_keys:
        raise ValueError("cap surface has no explicit ATM column")
    cap_rows = []
    for row in rows[2:]:
        if row[0] is None:
            continue
        expiry = norm_term(row[0])
        vols = {}
        for key, v in zip(strike_keys, row[1 : 1 + len(strike_keys)]):
            if v is None:
                raise ValueError(f"missing cap vol at expiry {expiry}, strike {key}")
            vols[key] = dec(v)
        cap_rows.append({"expiry": expiry, "vols": vols})

    # --- ATM swaption vol surface: banner row, header row, expiry rows ---------
    rows = read_rows(wb, "ATM_Swaption_Volatility")
    header = rows[1]
    assert str(header[0]).strip() == "Expiry"
    tenor_keys = [norm_term(c) for c in header[1:] if c is not None]
    swaption_rows = []
    for row in rows[2:]:
        if row[0] is None:
            continue
        expiry = norm_term(row[0])
        vols = {}
        for key, v in zip(tenor_keys, row[1 : 1 + len(tenor_keys)]):
            if v is None:
                raise ValueError(f"missing swaption vol at expiry {expiry}, tenor {key}")
            vols[key] = dec(v)
        swaption_rows.append({"expiry": expiry, "vols": vols})

    out = {
        "calibration_date": args.date,
        "currency": "USD",
        "discounting_index": "SOFR_OIS",
        "source_note": (
            f"Mid-market levels as of {args.date} from {os.path.basename(args.src)}. "
            "FHLB advance curve and derived Term LP (FHLB - SOFR, 1D pinned to 0) "
            "included. Cap strikes are absolute rates; the ATM column is a "
            "separate explicit ATM-strike quote."
        ),
        "generated_by": "research/convert_market_data.py",
        "curve_sofr_ois": {
            "day_count": "ACT/360",
            "fixed_payment_freq": "annual",
            "instruments": instruments,
        },
        "fhlb_curve": {
            "source": "FHLB advance offered rates, workbook FHLB_Curve sheet",
            "instruments": fhlb,
        },
        "tlp_nodes": tlp_nodes,
        "cap_vol_surface": {
            "vol_convention": "normal_bachelier",
            "underlying_index": "SOFR_3M",
            "strike_axis": "absolute_rate",
            "strikes": [k if k == "ATM" else float(k) for k in strike_keys],
            "rows": cap_rows,
        },
        "swaption_atm_vol_surface": {
            "vol_convention": "normal_bachelier",
            "underlying_index": "SOFR_swap",
            "rows": swaption_rows,
        },
    }

    with open(out_path, "w", encoding="utf-8", newline="\n") as f:
        json.dump(out, f, indent=1)
        f.write("\n")

    print(f"wrote {os.path.abspath(out_path)}")
    print(
        f"curve {len(instruments)} | fhlb {len(fhlb)} | tlp {len(tlp_nodes)} | "
        f"cap {len(cap_rows)}x{len(strike_keys)} | swaption {len(swaption_rows)}x{len(tenor_keys)}"
    )
    for n in tlp_nodes:
        print(f"  TLP {n['term']:>3} {n['spread']*1e4:7.2f} bp")


if __name__ == "__main__":
    main()
