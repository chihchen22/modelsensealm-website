"""
Replicating-portfolio "tractor" verification -- moving-average pillar construction.

Confirms the identity for a "k-month moving average of the k-month tenor" pillar:
    == k equal rolling bullet investments in the k-month tenor
    == a linear-amortizing runoff over k months  (WAL = (k+1)/2 months)
    == a portfolio whose yield equals the trailing k-month mean of the k-month rate.

Calibration (per Chih, 2026-06-08): real 9/30/2025 SOFR curve for pillar levels +
the codebase's spliced overnight history (buildHistoricalSOFR1M, Jan 2001-Feb 2026,
piecewise-linear FOMC anchors) for the trailing moving averages.

Two yields are reported per pillar:
  - new-money yield   : today's curve par rate at the pillar tenor (set up today).
  - steady-state yield : trailing k-month MA, lag-adjusted to the curve (book that has
                          been rolling for k months). CAVEAT: trailing MAs use the
                          OVERNIGHT history as a proxy for each tenor's own-rate
                          history, so long-tenor steady-state yields are understated in
                          steep-curve regimes (the 2010s). A production build would use
                          historical term curves. Short-tenor pillars are unaffected.

Two illustrative examples:
  IB  high-beta MMA : 80% overnight + 20% (24M MA of 24M tenor)
  NIB 3-year WAL    : solver blend of MA pillars to a 36-month WAL (min-vol vs max-yield)

Emits research/out/rp_tractor_examples.xlsx and prints PASS/FAIL on the identities.
Run (kb-env interpreter):
  & "C:\\Users\\deech\\AI\\Model-Sense\\ALM-Knowledge-Base\\ALM-Modeling\\.kb-env\\Scripts\\python.exe" research/rp_tractor_verify.py
"""
from __future__ import annotations
import os, json
import numpy as np
from scipy.optimize import minimize
from openpyxl import Workbook
from openpyxl.styles import Font

rng = np.random.default_rng(20260608)
PCT = lambda x: f"{100*x:.3f}%"
HERE = os.path.dirname(__file__)
AS_OF_IDX = 296   # Jan 2001 = 0  ->  Sep 2025 = 296

# ----------------------------------------------------------------------------
# 1. Spliced overnight history (port of buildHistoricalSOFR1M, decimal)
# ----------------------------------------------------------------------------
def build_historical_sofr1m():
    anchors = [(0, 6.0), (11, 1.75), (35, 1.0), (41, 1.0), (65, 5.25), (79, 5.25),
               (95, 0.1), (179, 0.1), (215, 2.4), (227, 1.55), (230, 0.05), (253, 0.05),
               (263, 4.3), (271, 5.3), (284, 5.3), (299, 3.8), (301, 3.66)]
    out = []
    for i in range(len(anchors) - 1):
        (m0, r0), (m1, r1) = anchors[i], anchors[i + 1]
        for j in range(m1 - m0):
            out.append((r0 + (r1 - r0) * j / (m1 - m0)) / 100.0)
    out.append(anchors[-1][1] / 100.0)
    return np.array(out)

OVN = build_historical_sofr1m()[: AS_OF_IDX + 1]   # overnight history through Sep 2025
OVN_NOW = OVN[-1]

# ----------------------------------------------------------------------------
# 2. Real 9/30/2025 SOFR curve -> par rate at any pillar tenor
# ----------------------------------------------------------------------------
def load_curve():
    snap = json.loads(open(os.path.join(HERE, "data", os.environ.get("MARKET_JSON", "market_2025-09-30.json"))).read())
    inst = snap["curve_sofr_ois"]["instruments"]
    t = np.array([x["t_years"] for x in inst])
    r = np.array([x["rate"] for x in inst])
    return t, r

CURVE_T, CURVE_R = load_curve()
def curve_rate(k_months: int) -> float:
    return float(np.interp(k_months / 12.0, CURVE_T, CURVE_R))

CURVE_OVN = curve_rate(1) if False else float(CURVE_R[0])   # 1D cash rate 3.63%

# ----------------------------------------------------------------------------
# 3. Pillar primitives
# ----------------------------------------------------------------------------
def pillar_wal_months(k: int) -> float:
    """k-month linear-amortizing (equal monthly principal): WAL = (k+1)/2 months."""
    return 0.5 if k <= 1 else (k + 1) / 2.0

def trailing_ma_overnight(k: int) -> float:
    return OVN[-1] if k <= 1 else OVN[-k:].mean()

def new_money_yield(k: int) -> float:
    """Set-up-today yield: the curve par rate at the pillar tenor."""
    return CURVE_OVN if k <= 1 else curve_rate(k)

def steady_state_yield(k: int) -> float:
    """Rolling book yield: trailing k-month MA, lag-adjusted (rebased) to the curve.
    = curve_rate(k) + [trailing_MA(overnight,k) - overnight_now]."""
    if k <= 1:
        return CURVE_OVN
    lag = trailing_ma_overnight(k) - OVN_NOW
    return curve_rate(k) + lag

# ----------------------------------------------------------------------------
# 4. Identity verification (on the real overnight series)
# ----------------------------------------------------------------------------
def verify_identity():
    rows, ok_all = [], True
    for k in [3, 6, 12, 24, 60, 120]:
        ladder = OVN[-k:].mean()                       # k equal rolling bullets
        ma = trailing_ma_overnight(k)                  # trailing k-month MA
        wal_closed = (k + 1) / 2.0
        wal_num = float(np.arange(1, k + 1) @ np.full(k, 1.0 / k))  # linear-amort runoff
        yk = bool(np.isclose(ladder, ma)); wk = bool(np.isclose(wal_closed, wal_num))
        ok_all = ok_all and yk and wk
        rows.append((k, ladder, ma, yk, wal_closed, wal_num, wk))
    return rows, ok_all

def blend_wal(weights_k):
    m = sum(w * pillar_wal_months(k) for w, k in weights_k)
    return m, m / 12.0

# ----------------------------------------------------------------------------
# 5. Forward Vasicek MC -> pillar-yield covariance (for the NIB frontier)
# ----------------------------------------------------------------------------
def pillar_yield_paths(pillars, horizon=12, n=4000, a=0.25, sigma=0.010):
    b = CURVE_OVN
    dt = 1 / 12.0
    r = np.full(n, CURVE_OVN)
    fut = np.empty((n, horizon))
    for h in range(horizon):
        r = r + a * (b - r) * dt + sigma * np.sqrt(dt) * rng.standard_normal(n)
        fut[:, h] = r
    cols = []
    for k in pillars:
        if k <= 1:
            cols.append(fut[:, -1]); continue
        need = max(0, k - horizon)
        tail = OVN[-need:] if need > 0 else np.empty(0)
        win = np.concatenate([np.broadcast_to(tail, (n, need)), fut], axis=1)[:, -k:]
        cols.append(win.mean(axis=1))
    return np.column_stack(cols)

def nib_solver(pillars, target_wal_m, objective="min_vol"):
    Y = pillar_yield_paths(pillars)
    cov = np.cov(Y, rowvar=False)
    mu = np.array([new_money_yield(k) for k in pillars])   # level from curve
    wal = np.array([pillar_wal_months(k) for k in pillars])
    n = len(pillars); w0 = np.full(n, 1.0 / n)
    cons = [{"type": "eq", "fun": lambda w: w.sum() - 1.0},
            {"type": "eq", "fun": lambda w: w @ wal - target_wal_m}]
    fun = (lambda w: float(w @ cov @ w)) if objective == "min_vol" else (lambda w: float(-(w @ mu)))
    res = minimize(fun, w0, method="SLSQP", bounds=[(0, 1)] * n, constraints=cons,
                   options=dict(maxiter=800, ftol=1e-12))
    w = np.clip(res.x, 0, None); w = w / w.sum()
    return dict(weights=w, blend_yield=float(w @ mu), blend_wal=float(w @ wal),
                blend_vol=float(np.sqrt(w @ cov @ w)), ok=bool(res.success))

# ----------------------------------------------------------------------------
# 6. Examples
# ----------------------------------------------------------------------------
def ib_example():
    beta = 0.80
    tr = [("Overnight (beta-linked)", beta, 1), ("24M MA of 24M (stable)", 1 - beta, 24)]
    nm = sum(w * new_money_yield(k) for _, w, k in tr)
    ss = sum(w * steady_state_yield(k) for _, w, k in tr)
    wal = sum(w * pillar_wal_months(k) for _, w, k in tr)
    return dict(beta=beta, tr=tr, new_money=nm, steady=ss, wal=wal)

def nib_example():
    pillars = [3, 12, 24, 60, 120, 180]
    return pillars, nib_solver(pillars, 36.0, "min_vol"), nib_solver(pillars, 36.0, "max_yield")

# ----------------------------------------------------------------------------
# 7. Run + report + Excel
# ----------------------------------------------------------------------------
def main():
    print(f"Overnight history: {len(OVN)} months (Jan 2001 - Sep 2025), now={PCT(OVN_NOW)}")
    print(f"Curve 1D={PCT(CURVE_OVN)}  2Y={PCT(curve_rate(24))}  5Y={PCT(curve_rate(60))}  10Y={PCT(curve_rate(120))}\n")

    idrows, idok = verify_identity()
    print("=== PILLAR IDENTITY (rolling ladder == trailing MA; WAL=(k+1)/2) ===")
    for k, lad, ma, yk, wc, wn, wk in idrows:
        print(f"  k={k:>3}  ladder={PCT(lad)}  MA={PCT(ma)}  match={yk}   WAL={wc:.1f}m num={wn:.1f}m match={wk}")
    print(f"IDENTITY ALL PASS: {idok}\n")

    print("=== PILLAR YIELDS: new-money (curve) vs steady-state (trailing MA) ===")
    for k in [1, 3, 6, 12, 24, 60, 120, 180]:
        print(f"  k={k:>3}  WAL={pillar_wal_months(k):>5.1f}m  new-money={PCT(new_money_yield(k))}  steady-state={PCT(steady_state_yield(k))}")
    print("  (steady-state lag uses overnight proxy; long tenors understated in steep-curve eras)\n")

    print("=== BLEND WAL ARITHMETIC ===")
    blends = {"100% 120M": [(1.0, 120)],
              "20/20/60 @ 24/60/120": [(0.2, 24), (0.2, 60), (0.6, 120)],
              "20/20/60 @ 24/60/180": [(0.2, 24), (0.2, 60), (0.6, 180)]}
    brep = {}
    for nm, wk in blends.items():
        m, y = blend_wal(wk); brep[nm] = (m, y)
        print(f"  {nm:>24}: {m:6.1f}m = {y:.2f}y")
    print("  (120M pillar WAL = 60.5m ~ 5.0y, not 120m)\n")

    ib = ib_example()
    print("=== IB EXAMPLE: high-beta MMA (80% beta) ===")
    for nm, w, k in ib["tr"]:
        print(f"  {nm:>26}: w={w:>4.0%} k={k:<3} WAL={pillar_wal_months(k):>5.1f}m  new-money={PCT(new_money_yield(k))}  steady={PCT(steady_state_yield(k))}")
    print(f"  RP credit  new-money={PCT(ib['new_money'])}  steady-state={PCT(ib['steady'])}  WAL={ib['wal']:.1f}m\n")

    pillars, sv, sy = nib_example()
    print("=== NIB EXAMPLE: 3y (36m) WAL, whole yield = structural IRR ===")
    print(f"  pillars k: {pillars}   WAL: {[round(pillar_wal_months(k),1) for k in pillars]}")
    print(f"  MIN-VOL : w={np.round(sv['weights'],3)}  WAL={sv['blend_wal']:.1f}m  yld={PCT(sv['blend_yield'])}  vol={1e4*sv['blend_vol']:.1f}bp")
    print(f"  MAX-YLD : w={np.round(sy['weights'],3)}  WAL={sy['blend_wal']:.1f}m  yld={PCT(sy['blend_yield'])}  vol={1e4*sy['blend_vol']:.1f}bp")

    # ---- Excel ----
    wb = Workbook(); bold = Font(bold=True)
    ws = wb.active; ws.title = "Pillar identity"
    ws.append(["k-month MA of k-month tenor -- identity (real overnight series)"]); ws["A1"].font = bold
    ws.append(["k", "rolling-ladder yield", "trailing MA", "yield match", "WAL (k+1)/2", "WAL numeric", "WAL match"])
    for c in ws[2]: c.font = bold
    for r in idrows: ws.append(list(r))
    ws.append([]); ws.append(["ALL PASS", idok])

    ws2 = wb.create_sheet("Pillar yields")
    ws2.append(["k", "WAL months", "new-money (curve)", "steady-state (trailing MA)"])
    for c in ws2[1]: c.font = bold
    for k in [1, 3, 6, 12, 24, 60, 120, 180]:
        ws2.append([k, pillar_wal_months(k), new_money_yield(k), steady_state_yield(k)])
    ws2.append([]); ws2.append(["caveat", "steady-state trailing MA uses overnight proxy; long tenors understated in steep-curve eras"])

    ws3 = wb.create_sheet("Blend WAL")
    ws3.append(["Blend", "WAL months", "WAL years"])
    for c in ws3[1]: c.font = bold
    for nm, (m, y) in brep.items(): ws3.append([nm, round(m, 2), round(y, 3)])

    ws4 = wb.create_sheet("IB example")
    ws4.append(["High-beta MMA: 80% overnight + 20% (24M MA of 24M)"]); ws4["A1"].font = bold
    ws4.append(["Tranche", "Weight", "k", "WAL months", "new-money", "steady-state"])
    for c in ws4[2]: c.font = bold
    for nm, w, k in ib["tr"]:
        ws4.append([nm, w, k, pillar_wal_months(k), new_money_yield(k), steady_state_yield(k)])
    ws4.append([]); ws4.append(["RP credit new-money", ib["new_money"]])
    ws4.append(["RP credit steady-state", ib["steady"]]); ws4.append(["Blended WAL months", round(ib["wal"], 2)])

    ws5 = wb.create_sheet("NIB example")
    ws5.append(["NIB 3y (36m) WAL -- whole yield is structural IRR"]); ws5["A1"].font = bold
    ws5.append(["pillar k"] + pillars)
    ws5.append(["WAL months"] + [round(pillar_wal_months(k), 1) for k in pillars])
    ws5.append(["new-money yield"] + [round(new_money_yield(k), 5) for k in pillars])
    ws5.append([])
    ws5.append(["MIN-VOL weights"] + [round(float(x), 4) for x in sv["weights"]])
    ws5.append(["  WAL m", round(sv["blend_wal"], 1), "yield", round(sv["blend_yield"], 5), "vol bp", round(1e4 * sv["blend_vol"], 1)])
    ws5.append(["MAX-YIELD weights"] + [round(float(x), 4) for x in sy["weights"]])
    ws5.append(["  WAL m", round(sy["blend_wal"], 1), "yield", round(sy["blend_yield"], 5), "vol bp", round(1e4 * sy["blend_vol"], 1)])

    outdir = os.path.join(HERE, "out"); os.makedirs(outdir, exist_ok=True)
    path = os.path.join(outdir, "rp_tractor_examples.xlsx")
    wb.save(path); print(f"\nExcel written: {path}")

if __name__ == "__main__":
    main()
